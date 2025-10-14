# parsers/excel_parser.py
# Excel HARA file parser with robust sheet/header detection
# Ported logic from fsc_hara_loader.py

import openpyxl
from typing import Optional, Dict, List
from cat.log import log
import sys
import os
import logging

# Add parent directory to path for imports
sys.path.append(os.path.dirname(os.path.dirname(__file__)))

from parsers.base_parser import BaseHARAParser, ParserError
from core.models import HaraData, SafetyGoal
from core.constants import (
    HARA_COLUMN_MAPPINGS, 
    DEFAULT_SAFE_STATE,
    DEFAULT_FTTI
)

try:
    import openpyxl
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    log.warning("openpyxl not available - Excel file reading will be disabled")


# Setup logging
log = logging.getLogger(__name__)


class ExcelHARAParser(BaseHARAParser):
    """
    Parser for Excel HARA files (.xlsx, .xls).
    Handles various column naming conventions, multiple sheets, and header rows not at row 1.
    
    Features:
    - Searches all sheets with priority (HARA Table, HARA, Safety Goals, etc.)
    - Finds headers in first 15 rows (handles title rows)
    - Flexible column name matching
    - Skips QM (non-safety-relevant) goals
    """
    
    # Sheet priority for HARA data (order matters)
    SHEET_PRIORITY = [
        'hara table',
        'hara',
        'hazard analysis',
        'safety goals',
        'fsc',
        'functional safety concept'
    ]
    
    def __init__(self, file_path: str, item_name: str):
        """
        Initialize Excel parser.
        
        Args:
            file_path: Path to Excel file
            item_name: Name of the item/system
        """
        super().__init__(item_name)
        self.file_path = file_path
        self.workbook = None
        self.worksheet = None
    
    def can_parse(self) -> bool:
        """Check if file exists and is Excel format"""
        if not os.path.exists(self.file_path):
            return False
        
        ext = os.path.splitext(self.file_path.lower())[1]
        return ext in ['.xlsx', '.xls']
    
    def parse(self) -> Optional[HaraData]:
        """
        Parse HARA data from Excel file.
        
        Returns:
            HaraData object if successful, None otherwise
        """
        try:
            log.warning(f"📖 Loading HARA from: {os.path.basename(self.file_path)}")
            
            # Load workbook
            self.workbook = openpyxl.load_workbook(self.file_path, data_only=True)
            log.info(f"📚 Available sheets: {self.workbook.sheetnames}")
            
            # Select worksheet with HARA data
            self.worksheet = self._select_worksheet()
            if not self.worksheet:
                log.warning("❌ Could not find worksheet with HARA data")
                return None
            
            log.info(f"📊 Selected sheet: {self.worksheet.title}")
            
            # Find and map column headers
            header_mapping = self._find_headers()
            if not header_mapping or 'safety_goal' not in header_mapping:
                log.warning("❌ Could not find 'Safety Goal' column in headers")
                return None
            
            log.info(f"✅ Columns found: {list(header_mapping.keys())}")
            
            # Extract safety goals
            goals = self._extract_goals(header_mapping)
            log.info(f"✅ Extracted {len(goals)} safety goals from Excel")
            
            if not goals:
                log.warning("⚠️ No safety goals extracted (all might be QM or filtered out)")
                return None
            
            # Create HaraData object
            return self._create_hara_data(
                goals,
                f"Excel: {os.path.basename(self.file_path)}, Sheet: {self.worksheet.title}"
            )
            
        except Exception as e:
            log.error(f"Error parsing Excel file: {e}")
            import traceback
            log.error(traceback.format_exc())
            raise ParserError(f"Error parsing Excel file: {e}")
        
        finally:
            if self.workbook:
                self.workbook.close()
    
    def _select_worksheet(self) -> Optional[object]:
        """
        Select the appropriate worksheet containing HARA data.
        
        Search strategy:
        1. Try priority sheet names (HARA Table, HARA, etc.)
        2. Fall back to active sheet
        
        Returns:
            Worksheet object or None
        """
        if not self.workbook:
            return None
        
        # Try priority sheets first
        for priority_name in self.SHEET_PRIORITY:
            for sheet_name in self.workbook.sheetnames:
                if priority_name in sheet_name.lower():
                    log.info(f"📊 Found priority sheet: {sheet_name}")
                    return self.workbook[sheet_name]
        
        # Fallback to active sheet
        log.info("📊 Using active sheet (no priority sheet found)")
        return self.workbook.active
    
    def _find_headers(self) -> Optional[Dict]:
        """
        Find and map column headers in worksheet.
        
        Features:
        - Searches first 15 rows for headers
        - Skips empty rows
        - Skips title rows (rows with <= 2 cells)
        - Handles multi-line text in cells
        - Case-insensitive matching
        - Flexible column name matching
        
        Returns:
            Dictionary mapping field names to column indices (1-based)
        """
        header_mapping = {}
        
        log.info("🔍 Searching for header row in first 15 rows...")
        
        # Search first 15 rows for headers (handles title rows)
        for row_idx in range(1, min(16, self.worksheet.max_row + 1)):
            row_values = [
                str(cell.value).lower().strip() if cell.value else ''
                for cell in self.worksheet[row_idx]
            ]
            
            # Skip completely empty rows
            if not any(row_values):
                log.debug(f"  Row {row_idx}: Empty, skipping")
                continue
            
            # Skip title rows (single merged cell or very few cells with content)
            non_empty = [v for v in row_values if v]
            if len(non_empty) <= 2:
                log.debug(f"  Row {row_idx}: Title row (only {len(non_empty)} cells), skipping")
                continue
            
            log.debug(f"  📋 Row {row_idx} values: {non_empty[:10]}")
            
            # Check if this looks like a header row
            # Must have at least 'safety goal' or 'goal' column
            has_goal = any(
                'safety goal' in val or val == 'goal' or 'safety goals' in val
                for val in row_values
            )
            
            if not has_goal:
                log.debug(f"  Row {row_idx}: No 'Safety Goal' column, skipping")
                continue
            
            # This is a header row - map all columns
            log.info(f"  ✅ Detected header row at row {row_idx}")
            header_mapping['header_row'] = row_idx
            
            for col_idx, val in enumerate(row_values, start=1):
                # Clean multi-line text
                val_clean = val.replace('\n', ' ').strip()
                
                # Safety Goal column
                if 'safety goal' in val_clean or val_clean == 'goal' or 'safety goals' in val_clean:
                    header_mapping['safety_goal'] = col_idx
                    log.info(f"    ✓ Safety Goal → Column {col_idx}")
                
                # ASIL column (critical!)
                elif 'asil' in val_clean and 'linked' not in val_clean and 'decomp' not in val_clean:
                    header_mapping['asil'] = col_idx
                    log.info(f"    ✓ ASIL → Column {col_idx}")
                
                # Safe State column
                elif 'safe state' in val_clean or val_clean == 'safe state':
                    header_mapping['safe_state'] = col_idx
                    log.info(f"    ✓ Safe State → Column {col_idx}")
                
                # FTTI column
                elif 'ftti' in val_clean or 'fault tolerant time' in val_clean:
                    header_mapping['ftti'] = col_idx
                    log.info(f"    ✓ FTTI → Column {col_idx}")
                
                # Severity column
                elif val_clean in ['s', 'severity', 'sev']:
                    header_mapping['severity'] = col_idx
                    log.info(f"    ✓ Severity → Column {col_idx}")
                
                # Exposure column
                elif val_clean in ['e', 'exposure', 'exp']:
                    header_mapping['exposure'] = col_idx
                    log.info(f"    ✓ Exposure → Column {col_idx}")
                
                # Controllability column
                elif val_clean in ['c', 'controllability', 'ctrl', 'control']:
                    header_mapping['controllability'] = col_idx
                    log.info(f"    ✓ Controllability → Column {col_idx}")
                
                # Hazard ID column
                elif 'hazard id' in val_clean or val_clean in ['id', 'hazard-id', 'haz id']:
                    header_mapping['hazard_id'] = col_idx
                    log.info(f"    ✓ Hazard ID → Column {col_idx}")
                
                # Hazardous Event column
                elif 'hazardous event' in val_clean or 'hazard event' in val_clean:
                    header_mapping['hazardous_event'] = col_idx
                    log.info(f"    ✓ Hazardous Event → Column {col_idx}")
                
                # Operational Situation column
                elif 'operational situation' in val_clean or 'operation' in val_clean:
                    header_mapping['operational_situation'] = col_idx
                    log.info(f"    ✓ Operational Situation → Column {col_idx}")
            
            # If we found headers, return them
            if 'safety_goal' in header_mapping:
                return header_mapping
        
        log.warning("❌ No valid header row found in first 15 rows")
        return None
    
    def _extract_goals(self, header_mapping: Dict) -> List[SafetyGoal]:
        """
        Extract safety goals from worksheet.
        
        Args:
            header_mapping: Column mapping dictionary
            
        Returns:
            List of SafetyGoal objects (QM goals are filtered out)
        """
        goals = []
        header_row = header_mapping.get('header_row', 1)
        sg_counter = 1
        
        # Check if ASIL column exists
        has_asil = 'asil' in header_mapping
        if not has_asil:
            log.warning("⚠️ ASIL column not found - will try to extract goals anyway")
        
        log.info(f"📊 Starting extraction from row {header_row + 1}")
        
        for row_idx in range(header_row + 1, self.worksheet.max_row + 1):
            row = self.worksheet[row_idx]
            
            # Get safety goal text
            sg_col = header_mapping.get('safety_goal')
            safety_goal_cell = row[sg_col - 1]
            safety_goal_text = str(safety_goal_cell.value).strip() if safety_goal_cell.value else ''
            
            # Skip empty rows or placeholder values
            if not safety_goal_text or safety_goal_text.lower() in ['none', 'n/a', '', 'null']:
                continue
            
            log.debug(f"  Row {row_idx}: Found goal - {safety_goal_text[:50]}...")
            
            # Get ASIL
            if has_asil:
                asil_col = header_mapping.get('asil')
                asil_cell = row[asil_col - 1]
                asil = str(asil_cell.value).strip() if asil_cell.value else 'QM'
                asil = self._normalize_asil(asil)
            else:
                # If no ASIL column, assume ASIL B (default)
                asil = 'B'
                log.warning(f"  Row {row_idx}: No ASIL column - defaulting to ASIL B")
            
            # Skip QM goals (not safety-relevant per ISO 26262)
            if asil == 'QM':
                log.debug(f"  Row {row_idx}: Skipping QM goal")
                continue
            
            # Extract all fields
            goal = SafetyGoal(
                id=self._generate_goal_id(sg_counter),
                description=self._clean_text(safety_goal_text),
                asil=asil,
                safe_state=self._get_cell_value(row, header_mapping.get('safe_state'), DEFAULT_SAFE_STATE),
                ftti=self._get_cell_value(row, header_mapping.get('ftti'), DEFAULT_FTTI),
                severity=self._get_cell_value(row, header_mapping.get('severity')),
                exposure=self._get_cell_value(row, header_mapping.get('exposure')),
                controllability=self._get_cell_value(row, header_mapping.get('controllability')),
                hazard_id=self._get_cell_value(row, header_mapping.get('hazard_id')),
                hazardous_event=self._get_cell_value(row, header_mapping.get('hazardous_event')),
                operational_situation=self._get_cell_value(row, header_mapping.get('operational_situation'))
            )
            
            goals.append(goal)
            log.info(f"  ✅ Extracted SG-{sg_counter:03d}: {safety_goal_text[:50]}... (ASIL {asil})")
            sg_counter += 1
        
        if len(goals) == 0:
            log.warning("⚠️ No goals extracted. Possible reasons:")
            log.warning("  - All goals are ASIL QM (not safety-relevant)")
            log.warning("  - Safety Goal column is empty")
            log.warning("  - Data starts after row 50 (not searched)")
        
        return goals
    
    def _get_cell_value(self, row, col_idx: Optional[int], default: str = "") -> str:
        """
        Safely get cell value from Excel row.
        
        Args:
            row: Excel row
            col_idx: Column index (1-based)
            default: Default value if cell is empty
            
        Returns:
            Cell value or default
        """
        if col_idx and col_idx <= len(row):
            val = row[col_idx - 1].value
            return self._clean_text(str(val)) if val else default
        return default
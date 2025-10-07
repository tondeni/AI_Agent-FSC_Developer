# utils.py
# Helper functions for FSC Developer Plugin
# Robust HARA parsing supporting multiple formats

from cat.log import log
import os
import re


def find_hara_data(cat, item_name):
    """
    Find HARA data from various sources with flexible format support.
    Priority: 1) Working memory, 2) hara_inputs folder, 3) Generated documents
    
    Supports multiple Excel formats and column naming conventions.
    """
    
    log.info(f"🔍 Finding HARA data for: {item_name}")
    
    # Try working memory first
    if "hara_table" in cat.working_memory:
        log.info("✅ Found HARA in working memory")
        return cat.working_memory["hara_table"]
    
    # Try hara_inputs folder
    plugin_folder = os.path.dirname(__file__)
    hara_folder = os.path.join(plugin_folder, "hara_inputs")
    
    log.info(f"📁 Looking in folder: {hara_folder}")
    
    if not os.path.exists(hara_folder):
        log.warning(f"⚠️ Folder doesn't exist: {hara_folder}")
        log.info(f"Creating hara_inputs folder: {hara_folder}")
        os.makedirs(hara_folder, exist_ok=True)
        return None
    
    # List all files in folder
    try:
        all_files = os.listdir(hara_folder)
        log.info(f"📋 Files in folder: {all_files}")
    except Exception as e:
        log.error(f"❌ Error listing folder: {e}")
        return None
    
    # Look for Excel files matching item name or any HARA file
    safe_name = "".join(c if c.isalnum() or c in "._- " else "_" 
                       for c in item_name).replace(" ", "_")
    
    log.info(f"🔍 Safe name for matching: {safe_name}")
    
    hara_files = []
    for filename in all_files:
        # Skip temporary Excel files (created when file is open)
        if filename.startswith('~$'):
            log.debug(f"⏭️ Skipping temp file: {filename}")
            continue
            
        if filename.endswith('.xlsx') or filename.endswith('.xls'):
            log.info(f"📄 Found Excel file: {filename}")
            # Prioritize files matching item name
            filename_lower = filename.lower()
            if safe_name.lower() in filename_lower or any(word.lower() in filename_lower for word in item_name.split()):
                log.info(f"✅ File matches item name: {filename}")
                hara_files.insert(0, filename)
            elif 'hara' in filename_lower:
                log.info(f"➕ File contains 'hara': {filename}")
                hara_files.append(filename)
    
    if not hara_files:
        log.warning(f"❌ No HARA Excel files found in {hara_folder}")
        return None
    
    log.info(f"📚 HARA files to try (in order): {hara_files}")
    
    # Try to read the first matching file
    for filename in hara_files:
        filepath = os.path.join(hara_folder, filename)
        log.info(f"📖 Attempting to read HARA file: {filename}")
        
        try:
            import openpyxl
            wb = openpyxl.load_workbook(filepath, data_only=True)
            log.info(f"✅ Workbook loaded, sheets: {wb.sheetnames}")
            
            # Try to find the HARA worksheet
            ws = find_hara_worksheet(wb)
            if not ws:
                log.warning(f"⚠️ No HARA worksheet found in {filename}")
                continue
            
            log.info(f"✅ Found HARA worksheet: {ws.title}")
            
            # Parse HARA data with flexible column mapping
            hara_data = parse_hara_worksheet(ws)
            
            if hara_data:
                log.info(f"✅ Successfully parsed {len(hara_data)} rows from {filename}")
                log.info(f"📊 Sample row keys: {list(hara_data[0].keys()) if hara_data else 'No data'}")
                return hara_data
            else:
                log.warning(f"⚠️ No valid data found in {filename}")
                
        except ImportError:
            log.error("❌ openpyxl not installed - cannot read Excel files")
            return None
        except Exception as e:
            log.error(f"❌ Error reading HARA file {filename}: {e}")
            import traceback
            log.error(traceback.format_exc())
            continue
    
    log.error("❌ Could not parse any HARA files")
    return None


def find_hara_worksheet(workbook):
    """
    Find the worksheet containing HARA data.
    Looks for sheets with HARA-related names or the first non-empty sheet.
    Priority: 'HARA Table' > 'HARA' > has required columns > other sheets
    """
    
    log.info(f"🔍 Looking for HARA worksheet in: {workbook.sheetnames}")
    
    # Priority 1: Exact matches for "HARA Table" or "HARA"
    priority_names = ['hara table', 'hara_table', 'hara']
    for priority_name in priority_names:
        for sheet_name in workbook.sheetnames:
            if sheet_name.lower() == priority_name:
                sheet = workbook[sheet_name]
                if has_required_hara_columns(sheet):
                    log.info(f"✅ Found priority HARA sheet: {sheet_name}")
                    return sheet
                else:
                    log.warning(f"⚠️ Sheet '{sheet_name}' doesn't have required HARA columns")
    
    # Priority 2: Sheets containing "HARA" or "Table" (more specific than just "safety")
    specific_keywords = ['hara', 'table', 'hazard', 'risk']
    for keyword in specific_keywords:
        for sheet_name in workbook.sheetnames:
            sheet_name_lower = sheet_name.lower()
            if keyword in sheet_name_lower:
                sheet = workbook[sheet_name]
                if has_required_hara_columns(sheet):
                    log.info(f"✅ Found HARA sheet by keyword '{keyword}': {sheet_name}")
                    return sheet
                else:
                    log.info(f"⚠️ Sheet '{sheet_name}' has keyword but missing required columns")
    
    # Priority 3: Any sheet with required HARA columns
    for sheet in workbook.worksheets:
        if has_required_hara_columns(sheet):
            log.info(f"✅ Found sheet with required columns: {sheet.title}")
            return sheet
    
    # Priority 4: Active sheet (last resort)
    if workbook.active and has_hara_data(workbook.active):
        log.warning(f"⚠️ Using active sheet (no better match): {workbook.active.title}")
        return workbook.active
    
    log.error("❌ No valid HARA worksheet found in workbook")
    return None


def has_hara_data(worksheet):
    """
    Check if worksheet contains HARA-like data.
    """
    
    if worksheet.max_row < 2:
        return False
    
    # Check first row for HARA-related headers
    first_row = [str(cell.value).lower() if cell.value else '' 
                 for cell in worksheet[1]]
    
    log.debug(f"🔍 Checking sheet '{worksheet.title}' headers: {first_row[:5]}...")
    
    hara_indicators = ['hazard', 'asil', 'safety goal', 'severity', 
                       'exposure', 'controllability', 'risk']
    
    has_data = any(indicator in ' '.join(first_row) for indicator in hara_indicators)
    
    if has_data:
        log.info(f"✅ Sheet '{worksheet.title}' has HARA indicators")
    
    return has_data


def has_required_hara_columns(worksheet):
    """
    Check if worksheet has the required HARA columns: ASIL and Safety Goal.
    More strict than has_hara_data() - ensures we can actually parse safety goals.
    Checks rows 1-10 to find the actual header row.
    """
    
    if worksheet.max_row < 2:
        log.debug(f"  Sheet '{worksheet.title}': Too few rows ({worksheet.max_row})")
        return False
    
    # Check rows 1-10 for headers (sometimes multiple title/empty rows)
    for row_idx in range(1, min(11, worksheet.max_row + 1)):
        headers = [str(cell.value).lower().strip() if cell.value else '' 
                   for cell in worksheet[row_idx]]
        
        # Skip completely empty rows
        non_empty = [h for h in headers if h]
        if len(non_empty) == 0:
            log.debug(f"  Row {row_idx}: Empty row, skipping")
            continue
        
        # Skip if this looks like a title row (single cell with content, rest empty)
        if len(non_empty) == 1:
            log.debug(f"  Row {row_idx}: Title row - {non_empty[0][:50]}")
            continue
        
        log.info(f"  📋 Sheet '{worksheet.title}' Row {row_idx} headers: {non_empty[:10]}")
        
        # Must have ASIL column
        has_asil = any('asil' in h for h in headers)
        
        # Must have Safety Goal column (be more flexible)
        has_sg = any(
            'safety goal' in h or 
            'safetygoal' in h or 
            'safety goals' in h or
            h == 'goal' or
            h == 'sg' for h in headers
        )
        
        # Alternative: Check for S, E, C columns (indicates HARA table structure)
        has_sec = all(h in headers for h in ['s', 'e', 'c'])
        
        log.info(f"  🔍 Row {row_idx}: has_asil={has_asil}, has_sg={has_sg}, has_SEC={has_sec}")
        
        # Accept if has ASIL and Safety Goal, OR if has S/E/C structure
        if (has_asil and has_sg) or has_sec:
            log.info(f"  ✅ Found valid headers in Row {row_idx}")
            return True
    
    log.warning(f"  ❌ No valid headers found in rows 1-10")
    log.warning(f"  💡 Please check if headers are beyond row 10 or in a different format")
    return False


def parse_hara_worksheet(worksheet):
    """
    Parse HARA worksheet with flexible column mapping.
    Handles various column naming conventions and formats.
    Automatically detects header row (checks rows 1-3).
    """
    
    log.info(f"📊 Parsing worksheet: {worksheet.title}")
    
    # Find the header row (could be row 1, 2, or 3)
    header_row_idx = find_header_row(worksheet)
    
    if not header_row_idx:
        log.error("❌ No header row found in worksheet")
        return None
    
    log.info(f"✅ Using header row: {header_row_idx}")
    
    # Get headers from detected row
    headers = []
    for cell in worksheet[header_row_idx]:
        header = str(cell.value).strip() if cell.value else ''
        headers.append(header)
    
    if not headers:
        log.error("❌ No headers found in HARA worksheet")
        return None
    
    log.info(f"📋 Found {len(headers)} headers: {[h for h in headers if h]}")
    
    # Create column mapping (flexible to handle different formats)
    column_map = create_column_mapping(headers)
    log.info(f"🗺️ Column mapping created: {len(column_map)} mappings")
    
    # Parse data rows (start from row after headers)
    hara_data = []
    for row_idx in range(header_row_idx + 1, worksheet.max_row + 1):
        row_data = {}
        
        for col_idx, header in enumerate(headers, 1):
            cell_value = worksheet.cell(row=row_idx, column=col_idx).value
            
            # Store with both original header and standardized key
            if header:
                row_data[header] = cell_value
            
            # Add standardized keys based on mapping
            std_key = column_map.get(header)
            if std_key:
                row_data[std_key] = cell_value
        
        # Only add row if it has meaningful data
        if has_meaningful_data(row_data):
            hara_data.append(row_data)
            log.debug(f"✅ Row {row_idx}: ASIL={row_data.get('ASIL')}, SG={str(row_data.get('Safety Goal', 'N/A'))[:50]}")
        else:
            log.debug(f"⚠️ Row {row_idx}: Skipped (no meaningful data)")
    
    log.info(f"✅ Parsed {len(hara_data)} valid rows from worksheet")
    
    if hara_data:
        log.info(f"📝 First row sample keys: {[k for k in hara_data[0].keys() if hara_data[0].get(k)][:10]}")
    
    return hara_data


def find_header_row(worksheet):
    """
    Find which row contains the actual column headers.
    Checks rows 1-10 for HARA-like headers.
    
    Returns:
        int: Row index (1-based) or None if not found
    """
    
    log.info(f"  🔍 Searching for header row in sheet '{worksheet.title}'")
    
    for row_idx in range(1, min(11, worksheet.max_row + 1)):
        headers = [str(cell.value).lower().strip() if cell.value else '' 
                   for cell in worksheet[row_idx]]
        
        # Skip completely empty rows
        non_empty = [h for h in headers if h]
        if len(non_empty) == 0:
            log.debug(f"    Row {row_idx}: Empty, skipping")
            continue
        
        # Skip if this looks like a title row (single cell with content)
        if len(non_empty) == 1:
            log.debug(f"    Row {row_idx}: Title - '{non_empty[0][:50]}'")
            continue
        
        # Log what we found
        log.info(f"    Row {row_idx}: {non_empty[:8]}")
        
        # Check if this row has HARA indicators
        has_hara_indicators = any(
            indicator in ' '.join(headers) 
            for indicator in ['asil', 'safety goal', 'hazard', 'severity', 'exposure', 'controllability']
        )
        
        if has_hara_indicators:
            log.info(f"  ✅ Row {row_idx} looks like headers!")
            return row_idx
    
    log.error(f"  ❌ No header row found in rows 1-10 of sheet '{worksheet.title}'")
    log.error(f"  💡 Total rows in sheet: {worksheet.max_row}")
    log.error(f"  💡 First 3 rows content:")
    for i in range(1, min(4, worksheet.max_row + 1)):
        row_preview = [str(cell.value)[:30] if cell.value else '' for cell in worksheet[i]]
        log.error(f"     Row {i}: {[c for c in row_preview if c][:5]}")
    
    return None


def create_column_mapping(headers):
    """
    Create flexible column mapping to handle various naming conventions.
    Maps actual column names to standardized keys.
    """
    
    column_map = {}
    
    for header in headers:
        if not header:
            continue
            
        header_lower = header.lower().strip()
        
        # Hazard ID
        if any(x in header_lower for x in ['hazard id', 'hazard_id', 'haz id', 'haz-id', 'id']):
            if 'hazard' in header_lower or 'haz' in header_lower:
                column_map[header] = 'Hazard ID'
                log.debug(f"  Map '{header}' -> 'Hazard ID'")
        
        # Function / Item
        elif any(x in header_lower for x in ['function', 'item', 'system', 'component']):
            column_map[header] = 'Function/Item'
            log.debug(f"  Map '{header}' -> 'Function/Item'")
        
        # Hazardous Event
        elif any(x in header_lower for x in ['hazardous event', 'hazard event', 'event']):
            column_map[header] = 'Hazardous Event'
            log.debug(f"  Map '{header}' -> 'Hazardous Event'")
        
        # Operational Situation
        elif any(x in header_lower for x in ['operational situation', 'operation', 'situation', 'scenario']):
            column_map[header] = 'Operational Situation'
            log.debug(f"  Map '{header}' -> 'Operational Situation'")
        
        # Severity
        elif header_lower in ['s', 'severity', 'sev', 's class']:
            column_map[header] = 'S'
            log.debug(f"  Map '{header}' -> 'S'")
        
        # Exposure
        elif header_lower in ['e', 'exposure', 'exp', 'e class']:
            column_map[header] = 'E'
            log.debug(f"  Map '{header}' -> 'E'")
        
        # Controllability
        elif header_lower in ['c', 'controllability', 'control', 'ctrl', 'c class']:
            column_map[header] = 'C'
            log.debug(f"  Map '{header}' -> 'C'")
        
        # ASIL
        elif any(x in header_lower for x in ['asil', 'asil rating', 'asil level']):
            column_map[header] = 'ASIL'
            log.debug(f"  Map '{header}' -> 'ASIL'")
        
        # Safety Goal
        elif any(x in header_lower for x in ['safety goal', 'safetygoal', 'sg', 'goal']):
            if 'safety' in header_lower:
                column_map[header] = 'Safety Goal'
                log.debug(f"  Map '{header}' -> 'Safety Goal'")
        
        # Safe State
        elif any(x in header_lower for x in ['safe state', 'safestate', 'ss']):
            column_map[header] = 'Safe State'
            log.debug(f"  Map '{header}' -> 'Safe State'")
        
        # FTTI
        elif any(x in header_lower for x in ['ftti', 'fault tolerant time', 'time interval']):
            column_map[header] = 'FTTI'
            log.debug(f"  Map '{header}' -> 'FTTI'")
    
    return column_map


def has_meaningful_data(row_data):
    """
    Check if row has meaningful data (not empty or just headers).
    """
    
    # Check for ASIL or Safety Goal
    asil = str(row_data.get('ASIL', '')).strip().upper()
    safety_goal = str(row_data.get('Safety Goal', '')).strip()
    
    # Clean ASIL value
    asil_clean = asil.replace('ASIL', '').replace('ASIL-', '').replace('-', '').strip()
    
    # Row is meaningful if it has a valid ASIL or substantial Safety Goal
    has_valid_asil = asil_clean in ['A', 'B', 'C', 'D', 'QM']
    has_valid_sg = len(safety_goal) > 5 and safety_goal.lower() not in ['safety goal', 'n/a', 'tbd']
    
    return has_valid_asil or has_valid_sg


def parse_safety_goals(hara_data):
    """
    Parse safety goals from HARA data.
    Robust parsing handling various formats and missing data.
    
    Per ISO 26262-3:2018, 7.4.2.2:
    At least one FSR shall be derived from each safety goal.
    """
    
    if not hara_data:
        log.error("❌ No HARA data to parse")
        return []
    
    log.info(f"🔍 Parsing {len(hara_data)} HARA rows for safety goals")
    
    safety_goals = []
    sg_counter = 1
    
    for idx, row in enumerate(hara_data, start=1):
        log.debug(f"📝 Processing row {idx}")
        
        # Extract and validate ASIL
        asil = extract_asil(row)
        log.debug(f"  ASIL: {asil}")
        
        if not asil:
            log.debug(f"  ⚠️ Row {idx}: No valid ASIL found")
            continue
            
        if asil == 'QM':
            log.debug(f"  ⚠️ Row {idx}: ASIL is QM, skipping")
            continue
        
        # Extract safety goal
        safety_goal_text = extract_safety_goal(row)
        log.debug(f"  Safety Goal: {safety_goal_text[:50] if safety_goal_text else 'None'}...")
        
        if not safety_goal_text:
            log.warning(f"⚠️ Row {idx} has ASIL {asil} but no safety goal - skipping")
            continue
        
        # Generate SG ID
        sg_id = f"SG-{sg_counter:03d}"
        
        # Extract other fields with fallbacks
        safe_state = extract_safe_state(row)
        ftti = extract_ftti(row)
        hazard_id = extract_hazard_id(row, sg_counter)
        
        # Create safety goal object
        safety_goal = {
            'id': sg_id,
            'description': safety_goal_text,
            'asil': asil,
            'safe_state': safe_state,
            'ftti': ftti,
            'hazard_id': hazard_id,
            'severity': extract_parameter(row, 'S', 'Severity'),
            'exposure': extract_parameter(row, 'E', 'Exposure'),
            'controllability': extract_parameter(row, 'C', 'Controllability'),
            'hazardous_event': extract_hazardous_event(row),
            'operational_situation': extract_operational_situation(row)
        }
        
        safety_goals.append(safety_goal)
        log.info(f"✅ Parsed {sg_id}: {asil} - {safety_goal_text[:60]}...")
        sg_counter += 1
    
    log.info(f"✅ Parsed {len(safety_goals)} safety goals from HARA data")
    
    if len(safety_goals) == 0:
        log.error("❌ No safety goals with ASIL A/B/C/D found in HARA")
        log.error("💡 Check that your HARA file has:")
        log.error("  1. An 'ASIL' column with values A, B, C, or D")
        log.error("  2. A 'Safety Goal' column with meaningful text")
    
    return safety_goals


def extract_asil(row):
    """
    Extract ASIL from row with flexible key matching.
    """
    
    possible_keys = ['ASIL', 'asil', 'ASIL Rating', 'ASIL Level']
    
    for key in possible_keys:
        if key in row and row[key]:
            asil = str(row[key]).strip().upper()
            asil = asil.replace('ASIL', '').replace('ASIL-', '').replace('-', '').strip()
            if asil in ['A', 'B', 'C', 'D', 'QM']:
                return asil
    
    return None


def extract_safety_goal(row):
    """
    Extract safety goal text with flexible key matching.
    """
    
    possible_keys = [
        'Safety Goal', 
        'SafetyGoal', 
        'Safety Goals',
        'Goal', 
        'SG',
        'Safety Requirement'
    ]
    
    for key in possible_keys:
        if key in row and row[key]:
            text = str(row[key]).strip()
            if len(text) > 5 and text.lower() not in ['safety goal', 'n/a', 'tbd', 'none']:
                return text
    
    return None


def extract_safe_state(row):
    """
    Extract safe state with fallback to default.
    Per ISO 26262-3:2018, 7.4.2.5
    """
    
    possible_keys = ['Safe State', 'SafeState', 'SS']
    
    for key in possible_keys:
        if key in row and row[key]:
            text = str(row[key]).strip()
            if text and text not in ['N/A', 'TBD', '-', 'None']:
                return text
    
    return "To be specified per ISO 26262-3:2018, 7.4.2.5"


def extract_ftti(row):
    """
    Extract FTTI with flexible format support.
    Per ISO 26262-3:2018, 7.4.2.4.b
    """
    
    possible_keys = [
        'FTTI', 
        'Fault Tolerant Time Interval', 
        'Time Interval',
        'Reaction Time',
        'Response Time'
    ]
    
    for key in possible_keys:
        if key in row and row[key]:
            ftti_value = str(row[key]).strip()
            if ftti_value and ftti_value not in ['N/A', 'TBD', '-', '', 'None']:
                return ftti_value
    
    return "To be determined per ISO 26262-3:2018, 7.4.2.4.b"


def extract_hazard_id(row, counter):
    """
    Extract hazard ID with fallback generation.
    """
    
    possible_keys = ['Hazard ID', 'Hazard_ID', 'HazardID', 'Haz ID', 'ID']
    
    for key in possible_keys:
        if key in row and row[key]:
            haz_id = str(row[key]).strip()
            if haz_id and haz_id not in ['N/A', 'TBD', '-', 'None']:
                return haz_id
    
    return f"H-{counter:03d}"


def extract_parameter(row, short_key, long_key):
    """
    Extract S, E, or C parameter.
    """
    
    if short_key in row and row[short_key]:
        value = str(row[short_key]).strip()
        if value and value not in ['N/A', 'TBD', '-', 'None']:
            return value
    
    if long_key in row and row[long_key]:
        value = str(row[long_key]).strip()
        if value and value not in ['N/A', 'TBD', '-', 'None']:
            return value
    
    return ''


def extract_hazardous_event(row):
    """
    Extract hazardous event description.
    """
    
    possible_keys = [
        'Hazardous Event',
        'Hazard Event',
        'Event',
        'Hazard',
        'Hazard Description'
    ]
    
    for key in possible_keys:
        if key in row and row[key]:
            text = str(row[key]).strip()
            if len(text) > 5:
                return text
    
    return 'Not specified'


def extract_operational_situation(row):
    """
    Extract operational situation.
    """
    
    possible_keys = [
        'Operational Situation',
        'Operating Situation',
        'Situation',
        'Scenario',
        'Operating Mode'
    ]
    
    for key in possible_keys:
        if key in row and row[key]:
            text = str(row[key]).strip()
            if text and text not in ['N/A', 'TBD', '-', 'None']:
                return text
    
    return 'General operation'


def validate_hara_data(safety_goals):
    """
    Validate parsed HARA data for ISO 26262-3 compliance.
    
    Per ISO 26262-3:2018, 7.3.1:
    HARA report shall be available as prerequisite.
    """
    
    if not safety_goals:
        return False, "No safety goals found with ASIL A/B/C/D"
    
    issues = []
    
    for sg in safety_goals:
        sg_id = sg.get('id', 'Unknown')
        
        if sg.get('asil') not in ['A', 'B', 'C', 'D']:
            issues.append(f"{sg_id}: Invalid ASIL '{sg.get('asil')}'")
        
        if not sg.get('description') or len(sg.get('description', '')) < 10:
            issues.append(f"{sg_id}: Safety goal description too short or missing")
        
        if 'To be specified' in sg.get('safe_state', ''):
            log.info(f"{sg_id}: Safe state to be specified (per ISO 26262-3:2018, 7.4.2.5)")
    
    if issues:
        log.warning(f"⚠️ HARA validation found {len(issues)} issues:")
        for issue in issues:
            log.warning(f"  - {issue}")
        
        critical_issues = [i for i in issues if 'Invalid ASIL' in i or 'too short' in i]
        return len(critical_issues) == 0, "; ".join(issues)
    
    log.info(f"✅ HARA validation passed: {len(safety_goals)} safety goals validated")
    return True, "All safety goals valid"


def format_safety_goals_summary(safety_goals):
    """
    Format safety goals for display.
    """
    
    if not safety_goals:
        return "No safety goals found"
    
    summary = f"Total Safety Goals: {len(safety_goals)}\n\n"
    
    # Group by ASIL
    by_asil = {}
    for sg in safety_goals:
        asil = sg.get('asil', 'Unknown')
        if asil not in by_asil:
            by_asil[asil] = []
        by_asil[asil].append(sg)
    
    # Display by ASIL level (D -> C -> B -> A)
    for asil in ['D', 'C', 'B', 'A']:
        if asil in by_asil:
            summary += f"\n**ASIL {asil}** ({len(by_asil[asil])} goals):\n"
            for sg in by_asil[asil][:5]:
                summary += f"- {sg['id']}: {sg['description'][:80]}...\n"
            if len(by_asil[asil]) > 5:
                summary += f"  ... and {len(by_asil[asil]) - 5} more\n"
    
    return summary